from flask import Flask, request, jsonify
from flask_socketio import SocketIO, emit
from flask_cors import CORS
import cv2
import numpy as np
import base64
import os
from deepface import DeepFace
import uuid
import time
import json
from openpyxl import Workbook, load_workbook



app = Flask(__name__)
CORS(app)
socketio = SocketIO(app, cors_allowed_origins="*")

# Directory to store registered face images
REGISTERED_FACES_DIR = "registered_faces"
os.makedirs(REGISTERED_FACES_DIR, exist_ok=True)

# Directory to store attendance records
ATTENDANCE_DIR = "attendance_records"
os.makedirs(ATTENDANCE_DIR, exist_ok=True)

# Directory to store face embeddings
EMBEDDINGS_DIR = "face_embeddings"
os.makedirs(EMBEDDINGS_DIR, exist_ok=True)

# In-memory database for users (in production, use a real database)
#users_db = {}
USERS_DB_FILE = "users_db.json"

# Load users_db from file if exists
if os.path.exists(USERS_DB_FILE):
    with open(USERS_DB_FILE, "r") as f:
        users_db = json.load(f)
else:
    users_db = {}


# Configuration for face recognition
FACE_MODEL = "Facenet"  # Using FaceNet model instead of VGG-Face
DETECTOR_BACKEND = "opencv"  # You can also try "retinaface", "mtcnn", "ssd", "dlib"
DISTANCE_METRIC = "cosine"  # Options: "cosine", "euclidean", "euclidean_l2"
RECOGNITION_THRESHOLD = 0.2  # Lower is more strict (for cosine similarity)

def base64_to_image(base64_string):
    """Convert base64 string to OpenCV image"""
    img_data = base64.b64decode(base64_string.split(',')[1])
    nparr = np.frombuffer(img_data, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    return img

def preprocess_face(img):
    """Preprocess face for better recognition results"""
    # Convert to grayscale (optional depending on model requirements)
    # gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # Apply histogram equalization for better lighting normalization
    # equalized = cv2.equalizeHist(gray)
    
    # Apply Gaussian blur to reduce noise
    blurred = cv2.GaussianBlur(img, (5, 5), 0)
    
    return blurred

@app.route('/api/register', methods=['POST'])
def register_face():
    data = request.json
    user_id = data.get('userId')
    name = data.get('name')
    base64_image = data.get('image')
    
    if not all([user_id, name, base64_image]):
        return jsonify({"success": False, "message": "Missing required data"}), 400
    
    try:
        # Convert base64 to image
        img = base64_to_image(base64_image)
        
        # Preprocess image for better face detection
        processed_img = preprocess_face(img)
        
        # Detect face using DeepFace
        face_objs = DeepFace.extract_faces(
            processed_img, 
            detector_backend=DETECTOR_BACKEND,
            enforce_detection=True,
            align=True
        )
        
        if not face_objs or len(face_objs) == 0:
            return jsonify({"success": False, "message": "No face detected. Please ensure good lighting and face the camera directly."}), 400
        
        # Save original image for reference
        filename = f"{user_id}_{uuid.uuid4()}.jpg"
        filepath = os.path.join(REGISTERED_FACES_DIR, filename)
        cv2.imwrite(filepath, img)
        
        # Extract and save face embedding - this is the key improvement with FaceNet
        embedding = DeepFace.represent(
            img_path=processed_img,
            model_name=FACE_MODEL,
            detector_backend=DETECTOR_BACKEND,
            enforce_detection=False,
            align=True
        )
        
        # Save embedding for future recognition
        embedding_file = os.path.join(EMBEDDINGS_DIR, f"{user_id}_{uuid.uuid4()}.npy")
        np.save(embedding_file, embedding)
        
        # Save user info
        users_db[user_id] = {
            "name": name,
            "face_path": filepath,
            "embedding_path": embedding_file,
            "registration_time": time.strftime("%Y-%m-%d %H:%M:%S")
        }

        # Save updated users_db to file
        with open(USERS_DB_FILE, "w") as f:
            json.dump(users_db, f, indent=4)

        return jsonify({
            "success": True, 
            "message": "Face registered successfully with FaceNet model",
            "userId": user_id
        })
        
    except Exception as e:
        print(f"Error registering face: {str(e)}")
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500

@socketio.on('recognize_face')
def recognize_face(data):
    base64_image = data.get('image')
    
    try:
        # Convert base64 to image
        img = base64_to_image(base64_image)
        
        # If no registered faces yet
        if not users_db:
            emit('recognition_result', {
                "success": False,
                "message": "No registered faces to compare with"
            })
            return
        
        # Preprocess image
        processed_img = preprocess_face(img)
        
        # Try to recognize the face using DeepFace
        try:
            # Use FaceNet model for face recognition
            result = DeepFace.find(
                img_path=processed_img, 
                db_path=REGISTERED_FACES_DIR, 
                detector_backend=DETECTOR_BACKEND,
                model_name=FACE_MODEL,
                distance_metric=DISTANCE_METRIC,
                enforce_detection=False,
                silent=True
            )
            
            if len(result) > 0 and len(result[0]) > 0:
                # Get the best match details
                matched_path = result[0]['identity'][0]
                match_distance = result[0]['distance'][0]
                
                # Check if the match is confident enough
                if match_distance <= RECOGNITION_THRESHOLD:
                    # Find the user who owns this face
                    recognized_user = None
                    for user_id, user_info in users_db.items():
                        if user_info["face_path"] in matched_path:
                            recognized_user = {
                                "userId": user_id,
                                "name": user_info["name"],
                                "confidence": round((1 - match_distance) * 100, 2)  # Convert distance to confidence percentage
                            }
                            break
                    
                    if recognized_user:
                        # Record attendance
                        attendance_file = os.path.join(ATTENDANCE_DIR, f"attendance_{time.strftime('%Y-%m-%d')}.xlsx")

                        already_marked = False

                        # Check if the Excel file exists
                        if os.path.exists(attendance_file):
                            # Load the workbook
                            wb = load_workbook(attendance_file)
                            ws = wb.active

                            # Check if user is already marked
                            for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
                                if row[0] == recognized_user['userId']:
                                    already_marked = True
                                    break
                        else:
                            # Create a new workbook and sheet
                            wb = Workbook()
                            ws = wb.active
                            # Add header row
                            ws.append(["User ID", "Name", "Date & Time", "Confidence"])

                        if already_marked:
                            emit('recognition_result', {
                                "success": True,
                                "user": recognized_user,
                                "message": f"Hi {recognized_user['name']}, you have already marked attendance today!"
                            })
                        else:
                            # Append new record
                            ws.append([
                                recognized_user['userId'],
                                recognized_user['name'],
                                time.strftime('%Y-%m-%d %H:%M:%S'),
                                recognized_user['confidence']
                            ])
                            wb.save(attendance_file)

                            emit('recognition_result', {
                                "success": True,
                                "user": recognized_user,
                                "message": f"Welcome, {recognized_user['name']}! (Confidence: {recognized_user['confidence']}%)"
                            })

                    else:
                        emit('recognition_result', {
                            "success": False,
                            "message": "Face recognized but user not found in database"
                        })
                else:
                    emit('recognition_result', {
                        "success": False,
                        "message": "Face detected but confidence too low for reliable identification"
                    })
            else:
                emit('recognition_result', {
                    "success": False,
                    "message": "Face not recognized"
                })
                
        except Exception as e:
            print(f"Recognition error: {str(e)}")
            emit('recognition_result', {
                "success": False,
                "message": f"Recognition error: {str(e)}"
            })
    
    except Exception as e:
        print(f"Processing error: {str(e)}")
        emit('recognition_result', {
            "success": False,
            "message": f"Processing error: {str(e)}"
        })

@app.route('/api/get-users', methods=['GET'])
def get_users():
    users_list = []
    for user_id, user_info in users_db.items():
        users_list.append({
            "userId": user_id,
            "name": user_info["name"],
            "registrationTime": user_info["registration_time"]
        })
    return jsonify({"users": users_list})


@app.route('/api/get-attendance', methods=['GET'])
def get_attendance():
    date = request.args.get('date', time.strftime('%Y-%m-%d'))
    attendance_file = os.path.join(ATTENDANCE_DIR, f"attendance_{date}.xlsx")
    
    attendance_list = []

    if os.path.exists(attendance_file):
        wb = load_workbook(attendance_file)
        ws = wb.active

        # Skip the header row (assumed to be the first row)
        for row in ws.iter_rows(min_row=2, values_only=True):
            user_id, name, timestamp, confidence = row
            attendance_list.append({
                "userId": user_id,
                "name": name,
                "timestamp": timestamp,
                "confidence": confidence
            })

    return jsonify({"attendance": attendance_list, "date": date})


@app.route('/api/model-info', methods=['GET'])
def model_info():
    """Return information about the face recognition model being used"""
    return jsonify({
        "model": FACE_MODEL,
        "detector": DETECTOR_BACKEND,
        "distance_metric": DISTANCE_METRIC,
        "threshold": RECOGNITION_THRESHOLD
    })

@app.route('/api/update-model-settings', methods=['POST'])
def update_model_settings():
    """Update face recognition model settings"""
    global FACE_MODEL, DETECTOR_BACKEND, DISTANCE_METRIC, RECOGNITION_THRESHOLD
    
    data = request.json
    if 'model' in data:
        FACE_MODEL = data['model']
    if 'detector' in data:
        DETECTOR_BACKEND = data['detector']
    if 'distance_metric' in data:
        DISTANCE_METRIC = data['distance_metric']
    if 'threshold' in data:
        RECOGNITION_THRESHOLD = float(data['threshold'])
        
    return jsonify({
        "success": True,
        "message": "Model settings updated",
        "settings": {
            "model": FACE_MODEL,
            "detector": DETECTOR_BACKEND,
            "distance_metric": DISTANCE_METRIC,
            "threshold": RECOGNITION_THRESHOLD
        }
    })

if __name__ == '__main__':
    socketio.run(app, host='0.0.0.0', port=5000, debug=True)